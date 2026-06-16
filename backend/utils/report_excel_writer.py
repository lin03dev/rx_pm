"""
Schema-driven Excel writer for configured report templates.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config.report_schema import get_report_template
from utils.excel_formatter import ExcelFormatter


class ReportExcelWriter:
    """Write report workbooks using schema_registry report_definitions."""

    def __init__(self) -> None:
        self._warn_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self._warn_font = Font(color="9C0006")
        self._good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self._good_font = Font(color="006100")
        self._pending_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self._pending_font = Font(color="9C6500")

    def save_report(
        self,
        output_path: Path | str,
        report_data: Dict[str, pd.DataFrame],
        sheet_names: Dict[str, str],
        report_id: Optional[str] = None,
    ) -> Path:
        path = Path(output_path)
        template = get_report_template(report_id) if report_id else None

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for sheet_key, df in report_data.items():
                display_name = sheet_names.get(sheet_key, sheet_key)[:31]
                payload = df.copy()
                if payload.empty:
                    payload = pd.DataFrame({"Message": ["No data available for this report"]})

                sheet_template = None
                if template:
                    sheet_template = (template.get("sheets") or {}).get(sheet_key)

                payload = self._order_columns(payload, sheet_template)
                payload.to_excel(writer, sheet_name=display_name, index=False)

                worksheet = writer.sheets.get(display_name)
                if worksheet is None:
                    continue

                if template and sheet_template:
                    self._apply_sheet_template(worksheet, payload, template, sheet_template)
                else:
                    ExcelFormatter.format_worksheet(worksheet)

        return path

    def _order_columns(self, df: pd.DataFrame, sheet_template: Optional[Dict[str, Any]]) -> pd.DataFrame:
        if not sheet_template:
            return df
        column_defs = sheet_template.get("columns") or []
        ordered = [col["name"] for col in column_defs if col.get("name") in df.columns]
        remaining = [name for name in df.columns if name not in ordered]
        if not ordered:
            return df
        return df[ordered + remaining]

    def _apply_sheet_template(
        self,
        worksheet,
        df: pd.DataFrame,
        template: Dict[str, Any],
        sheet_template: Dict[str, Any],
    ) -> None:
        defaults = template.get("defaults") or {}
        ExcelFormatter.format_worksheet(worksheet)

        if defaults.get("freeze_header", True):
            worksheet.freeze_panes = "A2"

        self._apply_column_layout(worksheet, df, sheet_template, defaults)
        self._apply_conditional_formats(worksheet, df, sheet_template)
        self._apply_summary_highlights(worksheet, df, sheet_template)
        self._apply_named_format(worksheet, df, sheet_template)

    def _apply_named_format(self, worksheet, df: pd.DataFrame, sheet_template: Dict[str, Any]) -> None:
        fmt = sheet_template.get("format") or sheet_template.get("role")
        if fmt == "attendance_matrix":
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if cell.value == "✓":
                        cell.fill = self._good_fill
                        cell.font = self._good_font
                    elif cell.value == "✗":
                        cell.fill = self._warn_fill
                        cell.font = self._warn_font
        elif fmt == "attendance_summary":
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        if cell.value >= 75:
                            cell.fill = self._good_fill
                        elif cell.value < 50:
                            cell.fill = self._warn_fill
                    elif isinstance(cell.value, str) and "%" in cell.value:
                        try:
                            val = float(cell.value.rstrip("%"))
                            if val >= 75:
                                cell.fill = self._good_fill
                            elif val < 50:
                                cell.fill = self._warn_fill
                        except ValueError:
                            pass
        elif fmt == "assignment_status":
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    val = cell.value.upper()
                    if "NOT SUBMITTED" in val or "REJECTED" in val:
                        cell.fill = self._warn_fill
                        cell.font = self._warn_font
                    elif "APPROVED" in val or "SUBMITTED" in val:
                        cell.fill = self._good_fill
                        cell.font = self._good_font
                    elif "REDO" in val:
                        cell.fill = self._pending_fill
                        cell.font = self._pending_font
        elif fmt == "participant_progress":
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if cell.value == 0 or cell.value == "0":
                        cell.fill = self._warn_fill
                        cell.font = self._warn_font
                    elif isinstance(cell.value, (int, float)) and cell.value > 0:
                        cell.fill = self._good_fill
                        cell.font = self._good_font
        elif fmt == "session_schedule":
            for i, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                if i % 2 == 0:
                    for cell in row:
                        cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    def _apply_column_layout(
        self,
        worksheet,
        df: pd.DataFrame,
        sheet_template: Dict[str, Any],
        defaults: Dict[str, Any],
    ) -> None:
        column_defs = {col["name"]: col for col in (sheet_template.get("columns") or []) if col.get("name")}
        max_width = int(sheet_template.get("max_column_width") or defaults.get("max_column_width") or 50)
        auto_width = sheet_template.get("auto_width", defaults.get("auto_width", True))

        for idx, column_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(idx)
            col_def = column_defs.get(column_name, {})
            width = col_def.get("width")
            if width:
                worksheet.column_dimensions[col_letter].width = min(int(width), max_width)
            elif auto_width:
                max_length = len(str(column_name))
                for row in range(2, min(len(df) + 2, 200)):
                    value = worksheet.cell(row=row, column=idx).value
                    if value is not None:
                        max_length = max(max_length, len(str(value)))
                worksheet.column_dimensions[col_letter].width = min(max_length + 2, max_width)

            align = col_def.get("align", "left")
            horizontal = "center" if align == "center" else "left"
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=idx)
                cell.alignment = Alignment(
                    horizontal=horizontal,
                    vertical="center",
                    wrap_text=bool((defaults.get("body") or {}).get("wrap_text", True)),
                )

    def _apply_conditional_formats(
        self,
        worksheet,
        df: pd.DataFrame,
        sheet_template: Dict[str, Any],
    ) -> None:
        for rule_group in sheet_template.get("conditional_formats") or []:
            column_name = rule_group.get("column")
            if column_name not in df.columns:
                continue
            col_idx = list(df.columns).index(column_name) + 1
            for row_idx in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                text = str(cell.value or "").strip()
                for rule in rule_group.get("rules") or []:
                    if self._rule_matches(text, rule):
                        cell.fill = PatternFill(
                            start_color=rule.get("fill", "FFFFFF"),
                            end_color=rule.get("fill", "FFFFFF"),
                            fill_type="solid",
                        )
                        cell.font = Font(color=rule.get("font", "000000"))
                        break

    def _apply_summary_highlights(
        self,
        worksheet,
        df: pd.DataFrame,
        sheet_template: Dict[str, Any],
    ) -> None:
        if sheet_template.get("layout") != "key_value":
            return

        metric_column = sheet_template.get("metric_column", "Metric")
        value_column = sheet_template.get("value_column", "Value")
        if metric_column not in df.columns or value_column not in df.columns:
            return

        value_idx = list(df.columns).index(value_column) + 1
        highlight_metrics = set(sheet_template.get("highlight_metrics") or [])

        for row_offset, metric in enumerate(df[metric_column].tolist(), start=2):
            if str(metric) in highlight_metrics:
                value_cell = worksheet.cell(row=row_offset, column=value_idx)
                value_cell.fill = self._warn_fill
                value_cell.font = self._warn_font

            if sheet_template.get("value_percent_highlight"):
                value_cell = worksheet.cell(row=row_offset, column=value_idx)
                text = str(value_cell.value or "")
                if "%" in text:
                    try:
                        number = float(text.rstrip("%").split()[0])
                        if number >= 75:
                            value_cell.fill = self._good_fill
                            value_cell.font = self._good_font
                        elif number < 50:
                            value_cell.fill = self._warn_fill
                            value_cell.font = self._warn_font
                    except ValueError:
                        pass

    @staticmethod
    def _rule_matches(text: str, rule: Dict[str, Any]) -> bool:
        if "equals" in rule:
            return text.lower() == str(rule["equals"]).lower()
        if "contains" in rule:
            return str(rule["contains"]).lower() in text.lower()
        return False


_writer: Optional[ReportExcelWriter] = None


def get_report_excel_writer() -> ReportExcelWriter:
    global _writer
    if _writer is None:
        _writer = ReportExcelWriter()
    return _writer
