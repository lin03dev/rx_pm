"""
Excel Formatter - Format Excel worksheets
"""

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelFormatter:
    """Utility class for formatting Excel worksheets"""
    
    @staticmethod
    def format_worksheet(worksheet, header_row=1):
        """Apply consistent formatting to worksheet"""
        if worksheet is None:
            return
            
        try:
            header_fill = PatternFill(start_color="1B4F72", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            if worksheet.max_row >= header_row:
                for cell in worksheet[header_row]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = thin_border
            
            for row in worksheet.iter_rows(min_row=header_row + 1, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = cell_alignment
                    cell.border = thin_border
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            worksheet.freeze_panes = f'A{header_row + 1}'
        except Exception as e:
            print(f"Warning: Could not format worksheet: {e}")