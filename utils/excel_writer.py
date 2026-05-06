"""
Excel Writer with Conditional Formatting
"""

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


class ConditionalExcelWriter:
    """Write Excel files with conditional formatting"""
    
    def __init__(self):
        self.green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        self.green_font = Font(color='006100')
        self.red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        self.red_font = Font(color='9C0006')
        self.yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        self.yellow_font = Font(color='9C6500')
        self.header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
        self.header_font = Font(color='FFFFFF', bold=True)
        self.light_green = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    
    def save_report(self, output_path: Path, report_data: dict, sheet_names: dict):
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_key, df in report_data.items():
                if not df.empty:
                    sheet_display = sheet_names.get(sheet_key, sheet_key)[:31]
                    df.to_excel(writer, sheet_name=sheet_display, index=False)
                    worksheet = writer.sheets[sheet_display]
                    self._apply_formatting(worksheet, sheet_display, df)
        return output_path
    
    def _apply_formatting(self, worksheet, sheet_title: str, df: pd.DataFrame):
        self._format_headers(worksheet)
        self._auto_size_columns(worksheet, df)
        
        title_lower = sheet_title.lower()
        
        if 'attendance matrix' in title_lower:
            self._format_attendance_matrix(worksheet)
        elif 'attendance summary' in title_lower:
            self._format_attendance_summary(worksheet)
        elif 'assignment status' in title_lower:
            self._format_assignment_status(worksheet)
        elif 'survey responses' in title_lower:
            self._format_survey_responses_simple(worksheet)
        elif 'participant progress' in title_lower:
            self._format_participant_progress(worksheet)
        elif 'summary statistics' in title_lower:
            self._format_summary_stats(worksheet)
        elif 'session schedule' in title_lower:
            self._format_session_schedule(worksheet)
    
    def _format_headers(self, worksheet):
        if worksheet.max_row >= 1:
            for cell in worksheet[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _auto_size_columns(self, worksheet, df):
        for idx, col in enumerate(df.columns, 1):
            max_length = len(str(col))
            col_letter = get_column_letter(idx)
            for row in range(2, min(len(df) + 2, 100)):
                cell_val = worksheet.cell(row=row, column=idx).value
                if cell_val:
                    max_length = max(max_length, len(str(cell_val)))
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    def _format_attendance_matrix(self, worksheet):
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value == '✓':
                    cell.fill = self.green_fill
                    cell.font = self.green_font
                elif cell.value == '✗':
                    cell.fill = self.red_fill
                    cell.font = self.red_font
    
    def _format_attendance_summary(self, worksheet):
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value and isinstance(cell.value, (int, float)):
                    if cell.value >= 75:
                        cell.fill = self.green_fill
                    elif cell.value < 50:
                        cell.fill = self.red_fill
                elif cell.value and isinstance(cell.value, str) and '%' in cell.value:
                    try:
                        val = float(cell.value.rstrip('%'))
                        if val >= 75:
                            cell.fill = self.green_fill
                        elif val < 50:
                            cell.fill = self.red_fill
                    except:
                        pass
    
    def _format_assignment_status(self, worksheet):
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.upper()
                    if 'NOT SUBMITTED' in val or '❌ NOT SUBMITTED' in val:
                        cell.fill = self.red_fill
                        cell.font = self.red_font
                    elif 'APPROVED' in val or '✅ APPROVED' in val:
                        cell.fill = self.green_fill
                        cell.font = self.green_font
                    elif 'SUBMITTED' in val or '📤 SUBMITTED' in val:
                        cell.fill = self.green_fill
                        cell.font = self.green_font
                    elif 'REDO' in val or '⚠️ REDO REQUIRED' in val:
                        cell.fill = self.yellow_fill
                        cell.font = self.yellow_font
    
    def _format_survey_responses_simple(self, worksheet):
        """Simply check if cells are empty and color accordingly"""
        # Find the selected_option and free_text_answer columns
        selected_col = None
        free_text_col = None
        
        for col_idx, cell in enumerate(worksheet[1], 1):
            cell_val = str(cell.value).lower() if cell.value else ''
            if 'selected_option' in cell_val:
                selected_col = col_idx
            elif 'free_text_answer' in cell_val:
                free_text_col = col_idx
        
        # Check both columns and color based on emptiness
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            # Check selected_option column
            if selected_col:
                cell = worksheet.cell(row=row[0].row, column=selected_col)
                # Check if empty (None, empty string, or just whitespace)
                is_empty = cell.value is None or (isinstance(cell.value, str) and not cell.value.strip())
                if is_empty:
                    # Also check if free_text has the answer (for text questions)
                    if free_text_col:
                        ft_cell = worksheet.cell(row=row[0].row, column=free_text_col)
                        ft_has_value = ft_cell.value and isinstance(ft_cell.value, str) and ft_cell.value.strip()
                        if not ft_has_value:
                            # Both columns empty - mark as NO RESPONSE
                            cell.fill = self.red_fill
                            cell.font = self.red_font
                            cell.value = '❌ NO RESPONSE'
                    else:
                        cell.fill = self.red_fill
                        cell.font = self.red_font
                        cell.value = '❌ NO RESPONSE'
                else:
                    # Has value - green
                    cell.fill = self.green_fill
                    cell.font = self.green_font
            
            # Check free_text_answer column
            if free_text_col:
                cell = worksheet.cell(row=row[0].row, column=free_text_col)
                is_empty = cell.value is None or (isinstance(cell.value, str) and not cell.value.strip())
                if is_empty:
                    # Check if selected_option has the answer
                    if selected_col:
                        so_cell = worksheet.cell(row=row[0].row, column=selected_col)
                        so_has_value = so_cell.value and isinstance(so_cell.value, str) and so_cell.value.strip()
                        if not so_has_value:
                            cell.fill = self.red_fill
                            cell.font = self.red_font
                            cell.value = '❌ NO RESPONSE'
                    else:
                        cell.fill = self.red_fill
                        cell.font = self.red_font
                        cell.value = '❌ NO RESPONSE'
                else:
                    # Has value - light green for text answers
                    cell.fill = self.light_green
                    cell.font = Font(color='006100')
    
    def _format_participant_progress(self, worksheet):
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value == 0 or cell.value == '0':
                    cell.fill = self.red_fill
                    cell.font = self.red_font
                elif cell.value and isinstance(cell.value, (int, float)) and cell.value > 0:
                    cell.fill = self.green_fill
                    cell.font = self.green_font
    
    def _format_summary_stats(self, worksheet):
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and '%' in cell.value:
                    try:
                        val = float(cell.value.rstrip('%').split()[0] if ' ' in cell.value else cell.value.rstrip('%'))
                        if val >= 75:
                            cell.fill = self.green_fill
                        elif val < 50:
                            cell.fill = self.red_fill
                    except:
                        pass
    
    def _format_session_schedule(self, worksheet):
        for i, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            if i % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')


_excel_writer = None

def get_excel_writer():
    global _excel_writer
    if _excel_writer is None:
        _excel_writer = ConditionalExcelWriter()
    return _excel_writer
