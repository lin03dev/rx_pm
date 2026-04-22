"""
Excel Template Generator - Creates Excel files with dropdowns and validation
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime

from config.excel_template_config import (
    ExcelTemplate,
    TemplateSheet,
    TemplateColumn,
    FieldType,
    ValidationType,
    DropdownOption,
    TemplatePurpose,
    get_excel_template_manager
)


class ExcelTemplateGenerator:
    """Generate Excel templates with dropdowns and validation"""
    
    def __init__(self, output_dir: str = "./output/templates"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_template(self, purpose: TemplatePurpose) -> Path:
        """Generate an Excel template for a specific purpose"""
        manager = get_excel_template_manager()
        template = manager.get_template(purpose)
        
        if not template:
            raise ValueError(f"Template not found for purpose: {purpose}")
        
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Create each sheet
        for sheet_config in template.sheets:
            ws = wb.create_sheet(title=sheet_config.sheet_name[:31])
            self._create_sheet(ws, sheet_config, template)
        
        # Add instructions sheet
        if template.instructions:
            self._add_instructions_sheet(wb, template)
        
        # Save the file
        file_path = self.output_dir / f"{template.name}.xlsx"
        wb.save(file_path)
        
        return file_path
    
    def _create_sheet(self, ws, sheet_config: TemplateSheet, template: ExcelTemplate):
        """Create a sheet with columns, dropdowns, and formatting"""
        
        # Sort columns by order
        columns = sorted(sheet_config.columns, key=lambda c: c.order)
        
        # Write headers
        headers = []
        for col in columns:
            header = col.display_name
            if col.required:
                header = f"{header} *"
            if col.description:
                header = f"{header}\n({col.description[:30]})"
            headers.append(header)
        
        ws.append(headers)
        
        # Apply header formatting
        self._format_header_row(ws, len(headers))
        
        # Add dropdown validations
        for col_idx, col in enumerate(columns, start=1):
            self._add_validation(ws, col, col_idx)
        
        # Add example row if provided
        if sheet_config.example_row:
            example_row = []
            for col in columns:
                value = sheet_config.example_row.get(col.field_name, "")
                example_row.append(value)
            ws.append(example_row)
            
            # Style example row
            example_row_num = ws.max_row
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=example_row_num, column=col_idx)
                cell.fill = PatternFill(start_color="E8F4FD", fill_type="solid")
                cell.font = Font(color="666666", italic=True)
        
        # Add notes row
        self._add_notes_row(ws, len(columns), sheet_config)
        
        # Auto-size columns
        self._auto_size_columns(ws, len(columns))
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _format_header_row(self, ws, column_count: int):
        """Format the header row"""
        header_fill = PatternFill(start_color="2E75B6", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx in range(1, column_count + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            
            ws.row_dimensions[1].height = 40
    
    def _add_validation(self, ws, col: TemplateColumn, col_idx: int):
        """Add data validation to a column"""
        if not col.validation:
            return
        
        col_letter = get_column_letter(col_idx)
        validation_range = f"{col_letter}2:{col_letter}1048576"
        
        if col.validation.validation_type == ValidationType.LIST and col.validation.options:
            dv = DataValidation(type="list", formula1=f'"{",".join([opt.value for opt in col.validation.options])}"')
            dv.error = col.validation.error_message
            dv.errorTitle = "Invalid Selection"
            ws.add_data_validation(dv)
            dv.add(validation_range)
        
        elif col.validation.validation_type == ValidationType.NUMBER_RANGE:
            dv = DataValidation(type="decimal", operator="between",
                formula1=col.validation.min_value, formula2=col.validation.max_value)
            dv.error = col.validation.error_message
            ws.add_data_validation(dv)
            dv.add(validation_range)
        
        elif col.validation.validation_type == ValidationType.TEXT_LENGTH:
            dv = DataValidation(type="textLength", operator="between",
                formula1=col.validation.min_length, formula2=col.validation.max_length)
            dv.error = col.validation.error_message
            ws.add_data_validation(dv)
            dv.add(validation_range)
    
    def _add_notes_row(self, ws, column_count: int, sheet_config: TemplateSheet):
        """Add a notes row with instructions"""
        notes_row_num = ws.max_row + 2
        
        ws.append([""] * column_count)
        
        notes_cell = ws.cell(row=notes_row_num + 1, column=1)
        notes_cell.value = f"📋 Notes: {sheet_config.description}"
        notes_cell.font = Font(color="666666", size=10, italic=True)
        
        required_fields = [c.display_name for c in sheet_config.columns if c.required]
        if required_fields:
            required_note = f"⚠️ Required fields: {', '.join(required_fields)}"
            ws.cell(row=notes_row_num + 2, column=1, value=required_note).font = Font(color="CC0000", size=10)
    
    def _auto_size_columns(self, ws, column_count: int):
        """Auto-size columns based on content"""
        for col_idx in range(1, column_count + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    length = len(str(cell_value))
                    if length > max_length:
                        max_length = min(length, 50)
            
            adjusted_width = max(max_length + 2, 15)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_instructions_sheet(self, wb: Workbook, template: ExcelTemplate):
        """Add an instructions sheet"""
        ws = wb.create_sheet(title="Instructions", index=0)
        
        ws['A1'] = f"📋 {template.display_name}"
        ws['A1'].font = Font(size=16, bold=True)
        
        ws['A3'] = "Description:"
        ws['A3'].font = Font(bold=True)
        ws['B3'] = template.description
        
        ws['A5'] = "Instructions:"
        ws['A5'].font = Font(bold=True)
        
        instructions_lines = template.instructions.strip().split('\n')
        for i, line in enumerate(instructions_lines, start=6):
            ws[f'A{i}'] = line.strip()
        
        ws['A100'] = f"Template Version: {template.version}"
        ws['A100'].font = Font(color="666666", size=9, italic=True)
        ws['B100'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['B100'].font = Font(color="666666", size=9, italic=True)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 30
    
    def generate_all_templates(self) -> List[Path]:
        """Generate all available templates"""
        manager = get_excel_template_manager()
        generated_files = []
        
        for purpose in TemplatePurpose:
            try:
                file_path = self.generate_template(purpose)
                generated_files.append(file_path)
                print(f"✅ Generated: {file_path.name}")
            except Exception as e:
                print(f"❌ Failed to generate {purpose.value}: {e}")
        
        return generated_files


class ExcelTemplateValidator:
    """Validate uploaded Excel files against templates"""
    
    def __init__(self):
        self.template_manager = get_excel_template_manager()
    
    def validate_upload(self, file_path: str, purpose: TemplatePurpose) -> Dict[str, Any]:
        """Validate an uploaded Excel file against a template"""
        template = self.template_manager.get_template(purpose)
        if not template:
            return {"valid": False, "error": f"Template not found for purpose: {purpose}"}
        
        try:
            excel_file = pd.ExcelFile(file_path)
            results = {"valid": True, "errors": [], "warnings": [], "sheets_validated": [], "rows_processed": 0}
            
            for sheet_config in template.sheets:
                if sheet_config.sheet_name not in excel_file.sheet_names:
                    results["errors"].append(f"Sheet '{sheet_config.sheet_name}' not found")
                    results["valid"] = False
                    continue
                
                df = pd.read_excel(file_path, sheet_name=sheet_config.sheet_name)
                sheet_result = self._validate_sheet(df, sheet_config, template)
                
                results["sheets_validated"].append({
                    "sheet_name": sheet_config.sheet_name,
                    "rows": len(df),
                    "errors": sheet_result["errors"],
                    "warnings": sheet_result["warnings"]
                })
                
                results["errors"].extend(sheet_result["errors"])
                results["warnings"].extend(sheet_result["warnings"])
                results["rows_processed"] += len(df)
            
            results["valid"] = len(results["errors"]) == 0
            return results
            
        except Exception as e:
            return {"valid": False, "error": str(e)}
    
    def _validate_sheet(self, df: pd.DataFrame, sheet_config: TemplateSheet, 
                        template: ExcelTemplate) -> Dict[str, Any]:
        """Validate a single sheet"""
        errors = []
        warnings = []
        
        expected_columns = [c.display_name for c in sheet_config.columns]
        actual_columns = list(df.columns)
        
        for col in sheet_config.columns:
            if col.required and col.display_name not in actual_columns:
                errors.append(f"Sheet '{sheet_config.sheet_name}': Required column '{col.display_name}' missing")
        
        for idx, row in df.iterrows():
            row_num = idx + 2
            
            for col in sheet_config.columns:
                if col.display_name not in df.columns:
                    continue
                
                value = row.get(col.display_name)
                
                if col.required and (pd.isna(value) or value == ""):
                    errors.append(f"Sheet '{sheet_config.sheet_name}', Row {row_num}: '{col.display_name}' is required")
                    continue
                
                if pd.isna(value) or value == "":
                    continue
                
                if col.field_type == FieldType.INTEGER:
                    try:
                        int(value)
                    except:
                        errors.append(f"Sheet '{sheet_config.sheet_name}', Row {row_num}: '{col.display_name}' must be an integer")
                
                elif col.field_type == FieldType.DATE:
                    try:
                        pd.to_datetime(value)
                    except:
                        errors.append(f"Sheet '{sheet_config.sheet_name}', Row {row_num}: '{col.display_name}' must be a valid date")
                
                elif col.field_type == FieldType.EMAIL:
                    if '@' not in str(value) or '.' not in str(value):
                        errors.append(f"Sheet '{sheet_config.sheet_name}', Row {row_num}: '{col.display_name}' must be a valid email")
                
                elif col.field_type == FieldType.JSON:
                    try:
                        import json
                        json.loads(str(value))
                    except:
                        errors.append(f"Sheet '{sheet_config.sheet_name}', Row {row_num}: '{col.display_name}' must be valid JSON")
        
        return {"errors": errors, "warnings": warnings}


# Singleton instances
_excel_template_generator = None
_excel_template_validator = None

def get_excel_template_generator() -> ExcelTemplateGenerator:
    global _excel_template_generator
    if _excel_template_generator is None:
        _excel_template_generator = ExcelTemplateGenerator()
    return _excel_template_generator

def get_excel_template_validator() -> ExcelTemplateValidator:
    global _excel_template_validator
    if _excel_template_validator is None:
        _excel_template_validator = ExcelTemplateValidator()
    return _excel_template_validator