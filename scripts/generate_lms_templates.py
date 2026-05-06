#!/usr/bin/env python3
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))
"""
Generate LMS Excel templates for data import
"""

import sys
import pandas as pd
from pathlib import Path
from datetime import datetime

sys.path.insert(0, '.')

from config.lms_templates_config import (
    LMSTemplateType, get_lms_template, get_all_lms_templates
)


def sanitize_sheet_name(name: str) -> str:
    """Sanitize sheet name for Excel"""
    invalid_chars = r'[\[\]\:\*\?/\\]'
    import re
    name = re.sub(invalid_chars, '_', name)
    return name[:31]


def generate_lms_template(template_type: LMSTemplateType, output_dir: Path) -> Path:
    """Generate a single LMS template"""
    template = get_lms_template(template_type)
    if not template:
        return None
    
    output_path = output_dir / f"{template.name}.xlsx"
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_config in template.sheets:
            sheet_name = sanitize_sheet_name(sheet_config['sheet_name'])
            
            # Create DataFrame with headers
            headers = [col['display'] for col in sheet_config['columns']]
            df = pd.DataFrame(columns=headers)
            
            # Add example row
            example_row = {}
            for col in sheet_config['columns']:
                example_row[col['display']] = col.get('example', '')
            df.loc[0] = example_row
            
            # Write to Excel
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get worksheet for formatting
            worksheet = writer.sheets[sheet_name]
            
            # Add comments/notes
            notes_row = len(df) + 2
            worksheet.cell(row=notes_row, column=1, value=f"📋 Instructions: {sheet_config.get('description', '')}")
            worksheet.cell(row=notes_row + 1, column=1, value=f"⚠️ Required fields are marked with *")
            
            # Add column help text
            help_row = notes_row + 3
            for idx, col in enumerate(sheet_config['columns'], 1):
                if col.get('help_text'):
                    worksheet.cell(row=help_row, column=idx, value=f"💡 {col['name']}: {col['help_text']}")
                    worksheet.cell(row=help_row, column=idx).font = Font(size=9, italic=True, color='666666')
    
    return output_path


def generate_all_lms_templates(output_dir: Path = None) -> List[Path]:
    """Generate all LMS templates"""
    if output_dir is None:
        output_dir = Path("./output/templates/LMS")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    generated = []
    
    print("="*60)
    print("LMS TEMPLATE GENERATOR")
    print("="*60)
    
    for template_type in LMSTemplateType:
        try:
            output_path = generate_lms_template(template_type, output_dir)
            if output_path:
                generated.append(output_path)
                print(f"✅ Generated: {output_path.name}")
        except Exception as e:
            print(f"❌ Failed to generate {template_type.value}: {e}")
    
    print(f"\n✅ Generated {len(generated)} templates")
    print(f"📁 Location: {output_dir}")
    
    return generated


def main():
    import argparse
    parser = argparse.ArgumentParser(description='Generate LMS Excel templates')
    parser.add_argument('--type', '-t', help='Template type to generate')
    parser.add_argument('--all', '-a', action='store_true', help='Generate all templates')
    parser.add_argument('--output', '-o', help='Output directory')
    
    args = parser.parse_args()
    
    output_dir = Path(args.output) if args.output else Path("./output/templates/LMS")
    
    if args.all:
        generate_all_lms_templates(output_dir)
    elif args.type:
        try:
            template_type = LMSTemplateType(args.type)
            output_path = generate_lms_template(template_type, output_dir)
            if output_path:
                print(f"✅ Generated: {output_path}")
        except ValueError:
            print(f"❌ Invalid template type: {args.type}")
            print(f"Available types: {[t.value for t in LMSTemplateType]}")
    else:
        # Generate all by default
        generate_all_lms_templates(output_dir)


if __name__ == '__main__':
    from openpyxl.styles import Font
    main()
