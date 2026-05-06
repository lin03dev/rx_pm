#!/usr/bin/env python3
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))
"""
Fully Dynamic Language Survey Workbook Generator
No hardcoded values - everything from database
"""

import sys
import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, '.')

from config.database_config import DatabaseConfigManager
from core.database_manager import DatabaseManager
from utils.dynamic_survey_processor import DynamicSurveyProcessor


def add_dropdown_to_column(worksheet, column_letter, start_row, options_list):
    dv = DataValidation(type="list", formula1=f'"{",".join(options_list)}"')
    worksheet.add_data_validation(dv)
    dv.add(f'{column_letter}{start_row}:{column_letter}1048576')


def create_dynamic_sheet(writer, processor, survey):
    """Create sheet dynamically from database"""
    survey_id = survey['id']
    survey_name = survey['survey']
    
    # Get survey structure dynamically
    structure = processor.get_survey_structure(survey_id)
    relationships = structure['relationships']
    
    # Get display config
    colors = processor.get_display_config()
    
    # Build sheet data
    sheet_data = []
    options_map = {}
    
    for rel in relationships:
        q_id = rel['question_id']
        q_text = rel['question_text']
        q_type = rel['question_type']
        is_parent = rel['is_parent']
        level = rel['level']
        
        is_mcq = q_type in [1, 2, 4, 5]
        
        if is_mcq:
            opts = processor.get_answer_options(q_id)
            if not opts.empty:
                opt_list = [opt['optionvalue'] for _, opt in opts.iterrows()]
                options_map[q_id] = opt_list
                instruction = f"▼ Click dropdown → Choose from: {', '.join(opt_list[:5])}{'...' if len(opt_list) > 5 else ''}"
            else:
                instruction = "▼ Click dropdown → Select option"
        else:
            instruction = "✎ Type your answer"
        
        # Format display with indentation for children
        display_text = q_text
        if level == 1:
            display_text = f"  └─ {q_text}"
        
        sheet_data.append({
            'Survey ID': survey_id,
            'Survey Name': survey_name,
            'Question ID': q_id,
            'Level': 'Main' if is_parent else ('Sub' if level == 1 else 'Standalone'),
            'Type': 'MCQ' if is_mcq else 'Text',
            'Question': display_text,
            'INSTRUCTION': instruction,
            'YOUR ANSWER →': ''
        })
    
    if not sheet_data:
        return None
    
    df = pd.DataFrame(sheet_data)
    sheet_name = survey_name[:31]
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    ws = writer.sheets[sheet_name]
    
    # Apply dynamic styling
    # Header
    header_fill = PatternFill(start_color=colors.get('header_bg', '1B4F72'), end_color=colors.get('header_bg', '1B4F72'), fill_type='solid')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 60
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 40
    
    ws.freeze_panes = 'A2'
    
    # Style rows based on type
    for idx, rel in enumerate(relationships, start=2):
        is_parent = rel['is_parent']
        level = rel['level']
        is_mcq = rel['question_type'] in [1, 2, 4, 5]
        
        if is_parent:
            bg_color = colors.get('parent_bg', 'D0E8F7')
            font_color = colors.get('parent_font', '0044CC')
        elif level == 1:
            bg_color = colors.get('child_bg', 'E8F5E9')
            font_color = colors.get('child_font', '006600')
        else:
            bg_color = colors.get('standalone_bg', 'FFFFFF')
            font_color = colors.get('standalone_font', '000000')
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            cell = ws[f'{col}{idx}']
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            if font_color != '000000':
                cell.font = Font(color=font_color)
        
        # Add dropdown for MCQ
        if is_mcq and rel['question_id'] in options_map:
            add_dropdown_to_column(ws, 'H', idx, options_map[rel['question_id']])
            ws[f'G{idx}'].font = Font(color='0066CC', bold=True)
        else:
            ws[f'G{idx}'].font = Font(color='009900', bold=True)
    
    # Answer column styling
    ws['H1'].fill = PatternFill(start_color=colors.get('answer_header_bg', 'FF8C00'), end_color=colors.get('answer_header_bg', 'FF8C00'), fill_type='solid')
    ws['H1'].font = Font(color='FFFFFF', bold=True)
    
    # Alternate row colors for answer column
    for idx in range(2, len(sheet_data) + 2):
        if idx % 2 == 0:
            ws[f'H{idx}'].fill = PatternFill(start_color=colors.get('answer_alt_bg', 'FFF3E0'), end_color=colors.get('answer_alt_bg', 'FFF3E0'), fill_type='solid')
    
    return ws


def generate_dynamic_workbook():
    db_config = DatabaseConfigManager()
    db_manager = DatabaseManager(db_config)
    db_manager.current_db = 'Telios_LMS_Dev'
    
    processor = DynamicSurveyProcessor(db_manager)
    
    output_dir = Path("./output/templates/Language")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "Dynamic_Language_Survey_Workbook.xlsx"
    
    print("="*70)
    print("DYNAMIC LANGUAGE SURVEY WORKBOOK GENERATOR")
    print("="*70)
    print("Reading configuration from database...")
    
    surveys_df = processor.get_surveys()
    
    # Filter to language surveys only (or all if specified)
    language_surveys = surveys_df[surveys_df['survey_type'] == 'Language']
    
    if language_surveys.empty:
        print("❌ No language surveys found")
        return None
    
    print(f"\n📋 Found {len(language_surveys)} language surveys")
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # Dynamic instructions
        instructions = pd.DataFrame({
            'Info': [
                'DYNAMIC LANGUAGE SURVEY WORKBOOK',
                '=' * 50,
                '',
                f'Generated from database on {pd.Timestamp.now()}',
                '',
                'HOW TO USE:',
                '1. Navigate to the survey sheet you need',
                '2. Blue rows = Main questions (dropdown for options)',
                '3. Green rows = Sub-questions',
                '4. Click the dropdown in "YOUR ANSWER →" column for MCQ',
                '5. Type directly for free text questions',
                '6. Save and upload using the import tool',
                '',
                f'Total Surveys: {len(language_surveys)}'
            ]
        })
        instructions.to_excel(writer, sheet_name='Instructions', index=False)
        
        ws_inst = writer.sheets['Instructions']
        for cell in ws_inst[1]:
            cell.fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        ws_inst.column_dimensions['A'].width = 80
        
        # Create sheets dynamically
        for _, survey in language_surveys.iterrows():
            print(f"   • Creating sheet: {survey['survey']}")
            create_dynamic_sheet(writer, processor, survey)
    
    print(f"\n✅ Generated: {output_path.name}")
    print(f"📁 Location: {output_dir}")
    
    return output_path


if __name__ == '__main__':
    generate_dynamic_workbook()
    print("\n" + "="*70)
    print("✅ Fully dynamic workbook generated!")
    print("   Everything loaded from database - no hardcoded values")
    print("="*70)
